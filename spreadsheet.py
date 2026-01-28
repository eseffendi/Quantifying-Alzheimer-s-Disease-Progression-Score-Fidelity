from openpyxl import load_workbook

import BERTScore
import ROUGE


INPUT_XLSX = "proxy_data.xlsx"
OUTPUT_XLSX = "output_scored.xlsx"

REFERENCE_COL = 1    # Column A
CANDIDATE_COL = 2    # Column B
START_OUTPUT_COL = 3 # Column C


HEADERS = [
    "BERT_Precision",
    "BERT_Recall",
    "BERT_F1",
    "ROUGE_1",
    "ROUGE_2",
    "ROUGE_L",
    "ROUGE_Lsum",
]


def ensure_headers(ws):
    """
    Write headers only if they are not already present.
    Assumes reference/candidate columns already exist.
    """
    for i, header in enumerate(HEADERS):
        col = START_OUTPUT_COL + i
        cell_value = ws.cell(row=1, column=col).value
        if cell_value != header:
            ws.cell(row=1, column=col, value=header)


def main():
    wb = load_workbook(INPUT_XLSX)
    ws = wb.active

    # ---- Ensure headers without overwriting data ----
    ensure_headers(ws)

    # ---- Iterate through all rows EXCEPT header row ----
    for row in range(2, ws.max_row + 1):
        reference = ws.cell(row=row, column=REFERENCE_COL).value
        candidate = ws.cell(row=row, column=CANDIDATE_COL).value

        if not reference or not candidate:
            continue

        # ---- BERTScore ----
        bert = BERTScore.get_BERT(candidate, reference)

        # ---- ROUGE ----
        rouge = ROUGE.get_ROUGE(candidate, reference)

        # ---- Write results ----
        ws.cell(row=row, column=START_OUTPUT_COL + 0, value=bert["precision"])
        ws.cell(row=row, column=START_OUTPUT_COL + 1, value=bert["recall"])
        ws.cell(row=row, column=START_OUTPUT_COL + 2, value=bert["f1"])

        ws.cell(row=row, column=START_OUTPUT_COL + 3, value=float(rouge["rouge1"]))
        ws.cell(row=row, column=START_OUTPUT_COL + 4, value=float(rouge["rouge2"]))
        ws.cell(row=row, column=START_OUTPUT_COL + 5, value=float(rouge["rougeL"]))
        ws.cell(row=row, column=START_OUTPUT_COL + 6, value=float(rouge["rougeLsum"]))

        print(f"Processed row {row}")

    wb.save(OUTPUT_XLSX)
    print(f"Saved results to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
